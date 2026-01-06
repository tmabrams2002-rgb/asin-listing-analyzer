import io
import re
import time
from dataclasses import dataclass
from typing import List, Optional, Tuple, Dict

import numpy as np
import pandas as pd
import requests
import streamlit as st
from PIL import Image

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo


# -----------------------------
# Password Gate (shared password, no accounts)
# -----------------------------
def password_gate():
    if "authed" not in st.session_state:
        st.session_state.authed = False

    if st.session_state.authed:
        return

    st.title("ASIN Listing Quality Analyzer")
    st.caption("Shared-password access only (no emails, no accounts).")

    pw = st.text_input("Password", type="password")
    if st.button("Enter"):
        if pw and pw == st.secrets.get("APP_PASSWORD", ""):
            st.session_state.authed = True
            st.rerun()
        else:
            st.error("Wrong password.")
    st.stop()


# -----------------------------
# Utilities
# -----------------------------
def norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(s).strip().lower()).strip()

def safe_str(x) -> str:
    return "" if pd.isna(x) else str(x)

def looks_like_url(x: str) -> bool:
    x = str(x)
    return x.startswith("http://") or x.startswith("https://")

def best_match_column(cols: List[str], candidates: List[str]) -> Optional[str]:
    cols_norm = {c: norm(c) for c in cols}
    best = None
    best_score = 0
    for c, cn in cols_norm.items():
        score = 0
        for cand in candidates:
            candn = norm(cand)
            for tok in candn.split():
                if tok and tok in cn:
                    score += 1
        if score > best_score:
            best_score = score
            best = c
    return best if best_score > 0 else None

def pick_bullet_columns(cols: List[str]) -> List[str]:
    cols_norm = [(c, norm(c)) for c in cols]
    candidates = []
    for c, cn in cols_norm:
        if ("bullet" in cn) or ("key product feature" in cn) or ("key feature" in cn):
            candidates.append(c)

    def sort_key(col):
        m = re.search(r"([1-9])", norm(col))
        return int(m.group(1)) if m else 99

    candidates.sort(key=sort_key)
    return candidates

def extract_bullets(row: pd.Series, bullet_cols: List[str]) -> List[str]:
    bullets = []
    for c in bullet_cols:
        v = safe_str(row.get(c, ""))
        if v.strip():
            bullets.append(v.strip())

    if not bullets:
        for c in row.index:
            cn = norm(c)
            if cn in ("bullet points", "bullets", "key product features", "key features"):
                v = safe_str(row.get(c, ""))
                if v.strip():
                    parts = [p.strip("• \t-") for p in re.split(r"[\r\n]+", v) if p.strip()]
                    bullets.extend([p for p in parts if p])

    return bullets[:5]

def find_image_url_columns(cols: List[str]) -> List[str]:
    # CLR variants differ; we’ll collect anything that looks like image url/link fields
    out = []
    for c in cols:
        cn = norm(c)
        if "image" in cn and ("url" in cn or "link" in cn):
            out.append(c)
        if cn in ("main image url", "mainimageurl", "main_image_url"):
            out.append(c)

    seen = set()
    dedup = []
    for c in out:
        if c not in seen:
            dedup.append(c)
            seen.add(c)
    return dedup


# -----------------------------
# Image Evaluation
# -----------------------------
@dataclass
class ImageAudit:
    asin: str
    slot: str
    url: str
    role: str
    score: int
    width: int
    height: int
    blur_flag: str
    low_res_flag: str
    background_flag: str
    notes: str
    fix: str

def download_image(url: str, timeout=15) -> Optional[Image.Image]:
    try:
        r = requests.get(url, timeout=timeout, headers={"User-Agent": "Mozilla/5.0"})
        r.raise_for_status()
        return Image.open(io.BytesIO(r.content)).convert("RGB")
    except Exception:
        return None

def laplacian_variance(gray: np.ndarray) -> float:
    gy = np.abs(np.diff(gray.astype(np.float32), axis=0)).mean()
    gx = np.abs(np.diff(gray.astype(np.float32), axis=1)).mean()
    return float(gx + gy)

def estimate_border_white_ratio(img: Image.Image) -> float:
    arr = np.asarray(img).astype(np.float32)
    h, w, _ = arr.shape
    b = 10
    border = np.concatenate([
        arr[0:b, :, :].reshape(-1, 3),
        arr[h-b:h, :, :].reshape(-1, 3),
        arr[:, 0:b, :].reshape(-1, 3),
        arr[:, w-b:w, :].reshape(-1, 3),
    ], axis=0)
    return float(np.mean((border[:, 0] > 240) & (border[:, 1] > 240) & (border[:, 2] > 240)))

def role_suggestion(slot: str, white_ratio: float) -> str:
    if slot == "Main" and white_ratio > 0.75:
        return "Main (white background)"
    if white_ratio > 0.75:
        return "Secondary (white background)"
    return "Lifestyle / Infographic"

def score_image(img: Image.Image, slot: str) -> Tuple[int, Dict[str, str], str, str]:
    arr = np.asarray(img)
    h, w = arr.shape[:2]
    longest = max(w, h)

    # Flags
    low_res = longest < 1000
    gray = np.dot(arr[..., :3], [0.299, 0.587, 0.114]).astype(np.uint8)
    blur_metric = laplacian_variance(gray)
    blur = blur_metric < 6.0  # heuristic
    white_ratio = estimate_border_white_ratio(img)
    bg_issue = (slot == "Main" and white_ratio < 0.55)

    score = 100
    if low_res: score -= 25
    if blur: score -= 25
    if bg_issue: score -= 20
    score = int(max(0, min(100, score)))

    notes = []
    fixes = []
    if low_res:
        notes.append("Low resolution (zoom risk).")
        fixes.append("Replace with 1600px+ image (longest side).")
    if blur:
        notes.append("Soft/blurred at thumbnail.")
        fixes.append("Use sharper source or reshoot with better focus/light.")
    if bg_issue:
        notes.append("Main image background not clean/white.")
        fixes.append("Use true white background; reduce shadows; center product.")
    if not notes:
        notes.append("Technically solid.")
        fixes.append("Keep; ensure image set includes lifestyle + infographic + dimensions.")

    flags = {
        "blur_flag": "Y" if blur else "N",
        "low_res_flag": "Y" if low_res else "N",
        "background_flag": "Y" if bg_issue else "N",
        "white_ratio": f"{white_ratio:.2f}",
        "blur_metric": f"{blur_metric:.2f}",
    }
    return score, flags, " ".join(notes), " ".join(fixes)


# -----------------------------
# Listing scoring (basic, stable rules)
# -----------------------------
def score_listing(title: str, bullets: List[str], desc: str, images: List[ImageAudit]) -> Tuple[int, int, int, List[str], List[str]]:
    issues, actions = [], []

    # Content (0–50)
    content = 50
    tlen = len(title.strip())
    if tlen < 90:
        content -= 10
        issues.append("Title may be too short / missing key attributes.")
        actions.append("Expand title with key attributes (size/material/use-case) without stuffing.")
    if tlen > 180:
        content -= 10
        issues.append("Title may be too long (truncation risk).")
        actions.append("Shorten title; front-load strongest keywords.")

    bcount = len([b for b in bullets if b.strip()])
    if bcount < 5:
        content -= (5 - bcount) * 5
        issues.append(f"Only {bcount} bullet(s) detected (aim for 5).")
        actions.append("Add missing bullets: benefits, specs, compatibility, what’s included, trust signals.")

    if bullets:
        avg_blen = int(np.mean([len(b) for b in bullets]))
        if avg_blen < 120:
            content -= 5
            issues.append("Bullets may be too short / light on benefits & specs.")
            actions.append("Rewrite bullets benefit-first with concrete specs.")

    if not desc.strip():
        content -= 5
        issues.append("Description appears missing/empty.")
        actions.append("Add concise description: use-case + what’s included + key specs.")

    content = max(0, min(50, content))

    # Images (0–25)
    img_score = 25
    if not images:
        img_score -= 20
        issues.append("No image URLs detected.")
        actions.append("Add main + 6+ supporting images (lifestyle, infographic, dimensions, packaging).")
    else:
        avg_img = int(np.mean([im.score for im in images]))
        if avg_img < 80:
            img_score -= 10
            issues.append("One or more images have quality/compliance risks.")
            actions.append("Replace low-res/blur/non-compliant images; ensure main image is clean.")
        if len(images) < 7:
            img_score -= 5
            issues.append("Fewer than 7 images detected.")
            actions.append("Add missing image types: dimensions infographic + in-use lifestyle + packaging.")
        main = next((im for im in images if im.slot == "Main"), None)
        if main and main.score < 80:
            img_score -= 5
            issues.append("Main image likely needs improvement.")
            actions.append("Replace main with crisp 1600px+ product-on-white, centered, no props/text.")

    img_score = max(0, min(25, img_score))

    # Signals (0–25) – CLR usually doesn’t contain reviews, keep neutral for now
    signals = 25

    overall = content + img_score + signals
    return overall, content, img_score, issues[:6], actions[:6]


# -----------------------------
# Excel export
# -----------------------------
def add_df_table(ws, df: pd.DataFrame, table_name: str):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        ws.append(row)
        if r_idx == 1:
            for c_idx in range(1, len(row) + 1):
                cell = ws.cell(row=1, column=c_idx)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    end_row = ws.max_row
    end_col = ws.max_column
    # simple A1 range; safe for <= 26 cols (fine for our output)
    ref = f"A1:{chr(64 + min(end_col, 26))}{end_row}"
    tab = Table(displayName=table_name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)
    ws.freeze_panes = "A2"

def build_excel(summary_df, content_df, image_df, action_df) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Listing Summary")
    add_df_table(ws, summary_df, "ListingSummary")

    ws = wb.create_sheet("Content Recommendations")
    add_df_table(ws, content_df, "ContentRecs")

    ws = wb.create_sheet("Image Audit")
    add_df_table(ws, image_df, "ImageAudit")

    ws = wb.create_sheet("Action Plan")
    add_df_table(ws, action_df, "ActionPlan")

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# -----------------------------
# Streamlit App
# -----------------------------
def parse_flat_file_rows(tsv_text: str) -> pd.DataFrame:
    rows = []
    for line in tsv_text.splitlines():
        if not line.strip():
            continue

        parts = line.split("\t")
        parts = [p.strip() for p in parts]

        # Find ASIN value by locating the literal "ASIN" token
        asin = ""
        for i, p in enumerate(parts):
            if p.upper() == "ASIN" and i + 1 < len(parts):
                asin = parts[i + 1].strip().upper()
                break

        # Title is typically the 2nd column in your sample
        title = parts[1] if len(parts) > 1 else ""

        # Grab image URLs anywhere in the row
        img_urls = [p for p in parts if p.startswith("http") and ".jpg" in p.lower()]

        if asin:
            rows.append({
                "ASIN": asin,
                "Title": title,
                "ImageURLs": img_urls
            })

    return pd.DataFrame(rows)
def main():
    password_gate()

    st.subheader("Input Mode")
    mode = st.radio(
        "Choose input mode",
        ["Upload CLR", "Paste ASINs + Upload CLR (analyze only those ASINs)", "Paste Flat File Rows (TSV)"],
        horizontal=False
    )

    # -----------------------------
    # Mode 3: Paste Flat File Rows
    # -----------------------------
    if mode == "Paste Flat File Rows (TSV)":
        flat_text = st.text_area(
            "Paste flat-file rows here (tab-delimited). One product per line.",
            height=240
        )

        st.subheader("Analysis Settings")
        eval_images = st.checkbox("Evaluate images individually", value=True)
        max_images_per_asin = st.slider("Max images per ASIN", 1, 12, 8)
        throttle = st.slider("Throttle between image downloads (seconds)", 0.0, 1.5, 0.2, 0.1)

        if not flat_text.strip():
            st.info("Paste your rows above to begin.")
            return

        df_flat = parse_flat_file_rows(flat_text)
        if df_flat.empty:
            st.error("No ASINs found. Make sure each row contains the word ASIN followed by the ASIN.")
            return

        st.caption(f"Parsed products: {len(df_flat):,}")

        if st.button("Analyze & Generate Excel"):
            summary_rows, content_rows, image_rows, action_rows = [], [], [], []
            prog = st.progress(0)
            total = len(df_flat)

            for i, (_, row) in enumerate(df_flat.iterrows(), start=1):
                asin = safe_str(row.get("ASIN", "")).upper().strip()
                title = safe_str(row.get("Title", ""))
                desc = ""  # not present in flat rows
                bullets = []  # not present in flat rows
                urls = list(row.get("ImageURLs", []))[:max_images_per_asin]

                audits: List[ImageAudit] = []
                if eval_images and urls:
                    for idx, url in enumerate(urls, start=1):
                        slot = "Main" if idx == 1 else f"Image{idx}"
                        img = download_image(url)
                        if img is None:
                            audits.append(ImageAudit(
                                asin=asin, slot=slot, url=url, role="Unknown",
                                score=0, width=0, height=0,
                                blur_flag="?", low_res_flag="?", background_flag="?",
                                notes="Could not download image (blocked/expired).",
                                fix="Verify URL or replace with accessible image link."
                            ))
                            continue

                        score, flags, notes, fix = score_image(img, slot)
                        role = role_suggestion(slot, float(flags["white_ratio"]))
                        audits.append(ImageAudit(
                            asin=asin, slot=slot, url=url, role=role,
                            score=score, width=img.size[0], height=img.size[1],
                            blur_flag=flags["blur_flag"],
                            low_res_flag=flags["low_res_flag"],
                            background_flag=flags["background_flag"],
                            notes=notes + f" (white_ratio={flags['white_ratio']}, blur_metric={flags['blur_metric']})",
                            fix=fix
                        ))
                        if throttle > 0:
                            time.sleep(throttle)

                overall, content_score, image_score, issues, actions = score_listing(title, bullets, desc, audits)

                summary_rows.append({
                    "ASIN": asin,
                    "Overall Score (0-100)": overall,
                    "Content Score (0-50)": content_score,
                    "Image Score (0-25)": image_score,
                    "Compliance Risk": "High" if overall < 70 else ("Med" if overall < 85 else "Low"),
                    "Biggest Issues": " | ".join(issues),
                    "Priority Actions": " | ".join(actions),
                })

                content_rows.append({
                    "ASIN": asin,
                    "Current Title": title,
                    "Detected Bullets (1-5)": "",
                    "Description": "",
                    "Recommendations (rules-based)": " | ".join(actions),
                    "Proposed Title (optional)": "",
                    "Proposed Bullet 1 (optional)": "",
                    "Proposed Bullet 2 (optional)": "",
                    "Proposed Bullet 3 (optional)": "",
                    "Proposed Bullet 4 (optional)": "",
                    "Proposed Bullet 5 (optional)": "",
                })

                for a in audits:
                    image_rows.append({
                        "ASIN": a.asin,
                        "Image Slot": a.slot,
                        "Image URL": a.url,
                        "Suggested Role": a.role,
                        "Score (0-100)": a.score,
                        "Width": a.width,
                        "Height": a.height,
                        "Blur Flag": a.blur_flag,
                        "Low Res Flag": a.low_res_flag,
                        "Background Flag": a.background_flag,
                        "Notes": a.notes,
                        "Recommended Fix": a.fix,
                    })

                for act in actions:
                    action_rows.append({
                        "ASIN": asin,
                        "Area": "Images" if "image" in act.lower() else "Content",
                        "Priority": "High" if overall < 70 else ("Med" if overall < 85 else "Low"),
                        "Estimated Impact": "High" if ("replace" in act.lower() or "main image" in act.lower()) else "Med",
                        "Action": act
                    })

                prog.progress(min(1.0, i / total))

            summary_df = pd.DataFrame(summary_rows).sort_values(by="Overall Score (0-100)")
            content_df = pd.DataFrame(content_rows)
            image_df = pd.DataFrame(image_rows) if image_rows else pd.DataFrame([{"Notes": "No images evaluated."}])
            action_df = pd.DataFrame(action_rows) if action_rows else pd.DataFrame([{"Action": "No actions generated."}])

            xlsx = build_excel(summary_df, content_df, image_df, action_df)

            st.success("Done. Download your Excel below.")
            st.download_button(
                "Download Excel (.xlsx)",
                data=xlsx,
                file_name="asin_listing_quality_export.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        return

    # -----------------------------
    # Mode 1/2: CLR Upload
    # -----------------------------
    uploaded = st.file_uploader("Upload Category Listing Report (XLSX/XLSM or CSV)", type=["xlsx", "xlsm", "csv"])

    asins = []
    if mode.startswith("Paste ASINs"):
        raw = st.text_area("Paste ASINs (one per line)", height=140)
        asins = [a.strip().upper() for a in raw.splitlines() if a.strip()]

    if not uploaded:
        st.info("Upload a CLR to begin.")
        return

    import openpyxl

    try:
        if uploaded.name.lower().endswith((".xlsx", ".xlsm")):
            wb = openpyxl.load_workbook(uploaded, read_only=True, keep_vba=True)

            sheet_name = st.selectbox(
                "Select sheet to analyze",
                wb.sheetnames,
                index=wb.sheetnames.index("Template") if "Template" in wb.sheetnames else 0
            )

            header_row = st.number_input(
                "Row number where column names start",
                min_value=1,
                max_value=50,
                value=4,
                step=1
            )

            df = pd.read_excel(uploaded, sheet_name=sheet_name, header=header_row - 1)
        else:
            df = pd.read_csv(uploaded)

    except Exception as e:
        st.error(f"Could not read file: {e}")
        return

    st.caption(f"Loaded rows: {len(df):,} | columns: {len(df.columns):,}")
    cols = list(df.columns)

    asin_guess = best_match_column(cols, ["asin", "product id", "product_id", "item id", "item_id"])
    if not asin_guess:
        st.error("Could not detect an ASIN column. Your CLR must include ASIN.")
        return

    df[asin_guess] = df[asin_guess].astype(str).str.upper().str.strip()

    if asins:
        df = df[df[asin_guess].isin(asins)].copy()
        st.caption(f"Filtered to pasted ASINs: {len(df):,}")
        if df.empty:
            st.warning("No matching ASINs found in the CLR.")
            return

    st.subheader("Column Mapping (auto-detected, adjust if needed)")

    title_guess = best_match_column(cols, ["item name", "title", "product name", "item_name"])
    desc_guess = best_match_column(cols, ["product description", "description", "product_description"])

    bullet_guess = pick_bullet_columns(cols)[:5]
    image_guess = find_image_url_columns(cols)

    left, right = st.columns(2)
    with left:
        asin_col = st.selectbox("ASIN column", cols, index=cols.index(asin_guess))
        title_col = st.selectbox(
            "Title column",
            ["(none)"] + cols,
            index=(["(none)"] + cols).index(title_guess) if title_guess in cols else 0
        )
        desc_col = st.selectbox(
            "Description column",
            ["(none)"] + cols,
            index=(["(none)"] + cols).index(desc_guess) if desc_guess in cols else 0
        )
    with right:
        bullet_cols = st.multiselect("Bullet columns (up to 5)", cols, default=bullet_guess)
        image_cols = st.multiselect("Image URL columns (main + additional)", cols, default=image_guess)

    st.subheader("Analysis Settings")
    eval_images = st.checkbox("Evaluate images individually", value=True)
    max_images_per_asin = st.slider("Max images per ASIN", 1, 12, 8)
    throttle = st.slider("Throttle between image downloads (seconds)", 0.0, 1.5, 0.2, 0.1)

    if st.button("Analyze & Generate Excel"):
        summary_rows, content_rows, image_rows, action_rows = [], [], [], []
        prog = st.progress(0)
        total = len(df)

        for i, (_, row) in enumerate(df.iterrows(), start=1):
            asin = safe_str(row.get(asin_col, "")).upper().strip()
            title = safe_str(row.get(title_col, "")) if title_col != "(none)" else ""
            desc = safe_str(row.get(desc_col, "")) if desc_col != "(none)" else ""
            bullets = extract_bullets(row, bullet_cols)

            urls = []
            for c in image_cols:
                u = safe_str(row.get(c, ""))
                if u and looks_like_url(u):
                    urls.append(u)
            urls = urls[:max_images_per_asin]

            audits: List[ImageAudit] = []
            if eval_images and urls:
                for idx, url in enumerate(urls, start=1):
                    slot = "Main" if idx == 1 else f"Image{idx}"
                    img = download_image(url)
                    if img is None:
                        audits.append(ImageAudit(
                            asin=asin, slot=slot, url=url, role="Unknown",
                            score=0, width=0, height=0,
                            blur_flag="?", low_res_flag="?", background_flag="?",
                            notes="Could not download image (blocked/expired).",
                            fix="Verify URL or replace with accessible image link."
                        ))
                        continue

                    score, flags, notes, fix = score_image(img, slot)
                    role = role_suggestion(slot, float(flags["white_ratio"]))
                    audits.append(ImageAudit(
                        asin=asin, slot=slot, url=url, role=role,
                        score=score, width=img.size[0], height=img.size[1],
                        blur_flag=flags["blur_flag"],
                        low_res_flag=flags["low_res_flag"],
                        background_flag=flags["background_flag"],
                        notes=notes + f" (white_ratio={flags['white_ratio']}, blur_metric={flags['blur_metric']})",
                        fix=fix
                    ))
                    if throttle > 0:
                        time.sleep(throttle)

            overall, content_score, image_score, issues, actions = score_listing(title, bullets, desc, audits)

            summary_rows.append({
                "ASIN": asin,
                "Overall Score (0-100)": overall,
                "Content Score (0-50)": content_score,
                "Image Score (0-25)": image_score,
                "Compliance Risk": "High" if overall < 70 else ("Med" if overall < 85 else "Low"),
                "Biggest Issues": " | ".join(issues),
                "Priority Actions": " | ".join(actions),
            })

            content_rows.append({
                "ASIN": asin,
                "Current Title": title,
                "Detected Bullets (1-5)": "\n".join([f"{j+1}. {b}" for j, b in enumerate(bullets)]),
                "Description": desc,
                "Recommendations (rules-based)": " | ".join(actions),
                "Proposed Title (optional)": "",
                "Proposed Bullet 1 (optional)": "",
                "Proposed Bullet 2 (optional)": "",
                "Proposed Bullet 3 (optional)": "",
                "Proposed Bullet 4 (optional)": "",
                "Proposed Bullet 5 (optional)": "",
            })

            for a in audits:
                image_rows.append({
                    "ASIN": a.asin,
                    "Image Slot": a.slot,
                    "Image URL": a.url,
                    "Suggested Role": a.role,
                    "Score (0-100)": a.score,
                    "Width": a.width,
                    "Height": a.height,
                    "Blur Flag": a.blur_flag,
                    "Low Res Flag": a.low_res_flag,
                    "Background Flag": a.background_flag,
                    "Notes": a.notes,
                    "Recommended Fix": a.fix,
                })

            for act in actions:
                action_rows.append({
                    "ASIN": asin,
                    "Area": "Images" if "image" in act.lower() else "Content",
                    "Priority": "High" if overall < 70 else ("Med" if overall < 85 else "Low"),
                    "Estimated Impact": "High" if ("replace" in act.lower() or "main image" in act.lower()) else "Med",
                    "Action": act
                })

            prog.progress(min(1.0, i / total))

        summary_df = pd.DataFrame(summary_rows).sort_values(by="Overall Score (0-100)")
        content_df = pd.DataFrame(content_rows)
        image_df = pd.DataFrame(image_rows) if image_rows else pd.DataFrame([{"Notes": "No images evaluated."}])



if __name__ == "__main__":
    main()
